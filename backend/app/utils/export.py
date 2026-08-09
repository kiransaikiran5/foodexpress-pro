import csv
import io
from typing import List, Dict, Any
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, PatternFill
from fastapi.responses import StreamingResponse

def export_csv(filename: str, headers: List[str], data: List[Dict[str, Any]]) -> StreamingResponse:
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    writer.writerows(data)
    output.seek(0)
    return StreamingResponse(
        iter([output.read()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}.csv"}
    )

def export_excel(filename: str, headers: List[str], data: List[Dict[str, Any]]) -> StreamingResponse:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename
    # header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
    # data rows
    for row_num, row in enumerate(data, 2):
        for col_num, key in enumerate(headers, 1):
            ws.cell(row=row_num, column=col_num, value=row.get(key, ""))
    # Auto-width
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"}
    )

def export_pdf(filename: str, headers: List[str], data: List[Dict[str, Any]]) -> StreamingResponse:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    header_style = ParagraphStyle(
        "HeaderStyle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10,
        alignment=1,
        textColor=colors.white,
        backColor=colors.HexColor("#4472C4"),
    )
    cell_style = ParagraphStyle(
        "CellStyle",
        parent=styles["Normal"],
        fontSize=8,
        alignment=1,
    )
    # Title
    elements = [Paragraph(filename, title_style), Spacer(1, 12)]
    # Build table data
    table_data = [headers] + [[str(row.get(h, "")) for h in headers] for row in data]
    # Style for header row
    table = Table(table_data)
    style_commands = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]
    table.setStyle(TableStyle(style_commands))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}.pdf"}
    )